#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
GMV SKU 自动化分析脚本（周期排序修复版）
"""

from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Iterable, List, Dict, Tuple

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference, LineChart
from openpyxl.chart.label import DataLabelList
from openpyxl.drawing.colors import ColorChoice

# 关闭pandas未来警告，彻底消除警告信息
pd.set_option('future.no_silent_downcasting', True)

# 核心字段定义
ORDER_COL = "Order ID"
ORDER_STATUS_COL = "order status"
ITEM_ID_COL = "Order item ID"
ITEM_STATUS_COL = "order item status"
SKU_COL = "contribution sku"
QTY_COL = "quantity purchased"
PRODUCT_NAME_COL = "product name"
VARIATION_COL = "variation"
STORE_COL = "店铺"
PERIOD_COL = "周期"

# 金额字段定义
BASE_COL = "Base price total"
RETAIL_COL = "Retail price total after discounts(tax excl.)"
SHIP_COL = "Shipping total(tax excl.)"
REFUND_COL = "Product refund"
MONEY_COLS = [BASE_COL, RETAIL_COL, SHIP_COL, REFUND_COL]

# Excel样式定义
HEADER_FILL = "1F4E78"
TITLE_FILL = "0F766E"
WHITE = "FFFFFF"
THIN = Side(style="thin", color="D9E2F3")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

# 图表配色定义（专业商务风，对比清晰）
SALES_BAR_COLOR = "2F5597"  # 销量柱形：深蓝色
GMV_LINE_COLOR = "00B050"   # GMV折线：深绿色
LABEL_WHITE = "FFFFFF"       # 标签白色字体
LABEL_BLACK = "1F4E78"       # 标签黑色字体


# -------------------------- 核心修复：通用周期排序函数，全脚本统一使用 --------------------------
def period_sort_key(period_str: str) -> Tuple[int, int, int, int]:
    """
    周期文本排序key函数，按时间数字排序，彻底解决字符串排序错乱问题
    输入："5.8-5.14" → 输出：(5, 8, 5, 14)
    输入："5.15-5.21" → 输出：(5, 15, 5, 21)
    排序规则：起始月 → 起始日 → 结束月 → 结束日
    """
    # 提取周期中的所有数字
    nums = [int(n) for n in re.findall(r"\d+", str(period_str))]
    # 补全4位数字，处理空值、异常格式，避免排序报错
    while len(nums) < 4:
        nums.append(999)
    # 返回前4位，用于排序
    return (nums[0], nums[1], nums[2], nums[3])
# -----------------------------------------------------------------------------------------


def money_to_float(value) -> float:
    """去掉 MX$、逗号、制表符等字符，转成 float。空值按 0 处理。"""
    if pd.isna(value):
        return 0.0
    s = str(value).strip().replace("\t", "").replace(",", "")
    if s == "" or s.lower() in {"nan", "none", "null", "-"}:
        return 0.0
    s = re.sub(r"[^0-9.\-]", "", s)
    if s in {"", "-", ".", "-."}:
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def clean_text(value) -> str:
    if pd.isna(value):
        return ""
    return str(value).replace("\t", "").strip()


def infer_store_from_filename(path: Path) -> str:
    """【修复版】正确提取店铺名，避免把日期数字包含进去，统一多周期店铺名"""
    name = path.stem
    # 提取文件名开头的纯字母/中文部分作为店铺名，核心修复点
    match = re.match(r"^([a-zA-Z\u4e00-\u9fa5]+)", name)
    if match:
        return match.group(1).strip()
    # 兜底逻辑
    return re.split(r"[-_ ]", name, maxsplit=1)[0].strip() or "店铺"


def infer_period_from_filename(path: Path) -> str:
    """从文件名提取统计周期，支持x.x-x.x格式"""
    m = re.search(r"(\d{1,2}\.\d{1,2})\s*-\s*(\d{1,2}\.\d{1,2})", path.name)
    return f"{m.group(1)}-{m.group(2)}" if m else path.stem


def collect_csv_files(inputs: Iterable[str]) -> List[Path]:
    files: List[Path] = []
    for item in inputs:
        p = Path(item)
        if p.is_dir():
            files.extend(sorted(p.glob("*.csv")))
        elif p.is_file() and p.suffix.lower() == ".csv":
            files.append(p)
        else:
            raise FileNotFoundError(f"找不到 CSV 文件或文件夹：{item}")
    if not files:
        raise FileNotFoundError("没有找到任何 CSV 文件。")
    return files


def read_csv_smart(path: Path) -> pd.DataFrame:
    for enc in ["utf-8-sig", "utf-8", "gbk", "latin1"]:
        try:
            return pd.read_csv(path, encoding=enc, dtype=str)
        except UnicodeDecodeError:
            continue
    return pd.read_csv(path, dtype=str)


def load_one_csv(path: Path) -> pd.DataFrame:
    df = read_csv_smart(path)
    df["来源文件"] = path.name
    df[STORE_COL] = infer_store_from_filename(path)
    df[PERIOD_COL] = infer_period_from_filename(path)

    # 自动适配SKU字段
    if SKU_COL not in df.columns:
        for fallback in ["seller sku", "Seller SKU", "SKU ID", VARIATION_COL]:
            if fallback in df.columns:
                df[SKU_COL] = df[fallback]
                break
    if SKU_COL not in df.columns:
        raise KeyError(f"{path.name} 缺少 SKU 字段：{SKU_COL}")

    # 补全缺失字段
    for col in [ORDER_COL, ORDER_STATUS_COL, ITEM_ID_COL, ITEM_STATUS_COL, SKU_COL, PRODUCT_NAME_COL, VARIATION_COL]:
        if col not in df.columns:
            df[col] = ""
        df[col] = df[col].map(clean_text)

    if QTY_COL not in df.columns:
        df[QTY_COL] = 1

    return df


def remove_canceled_orders(df: pd.DataFrame) -> Tuple[pd.DataFrame, int]:
    possible_status_cols = [ORDER_STATUS_COL, ITEM_STATUS_COL, "Order status", "Order item status", "status", "Status"]
    status_cols = [c for c in possible_status_cols if c in df.columns]
    if not status_cols:
        return df.copy(), 0
    status_text = df[status_cols].fillna("").astype(str).agg(" ".join, axis=1).str.lower()
    canceled_mask = status_text.str.contains(r"cancel|canceled|cancelled|取消", regex=True, na=False)
    return df.loc[~canceled_mask].copy(), int(canceled_mask.sum())


def calculate_detail(df: pd.DataFrame) -> pd.DataFrame:
    out = df.copy()

    for col in MONEY_COLS:
        if col not in out.columns:
            out[col] = 0
        out[col] = out[col].map(money_to_float)

    out[QTY_COL] = out[QTY_COL].map(money_to_float)
    out[ORDER_COL] = out[ORDER_COL].map(clean_text)
    out[SKU_COL] = out[SKU_COL].map(clean_text)
    out[PRODUCT_NAME_COL] = out[PRODUCT_NAME_COL].map(clean_text)
    out[VARIATION_COL] = out[VARIATION_COL].map(clean_text)

    retail = out[RETAIL_COL]
    shipping = out[SHIP_COL]
    base = out[BASE_COL]
    refund = out[REFUND_COL]

    # 严格按Temu墨西哥站结算规则计算GMV
    out["返还VAT（应收）"] = (retail + shipping) * 0.08
    out["ISR预扣（应扣）"] = -((retail + shipping) * 0.025)
    out["尾程预估（应扣）"] = 32.0
    out["GMV"] = base + shipping + out["返还VAT（应收）"] + out["ISR预扣（应扣）"] - out["尾程预估（应扣）"] - refund
    return out


def first_non_empty(series: pd.Series) -> str:
    for v in series:
        s = clean_text(v)
        if s:
            return s
    return ""


def build_summaries(detail: pd.DataFrame):
    # 【修复】周期按时间数字排序，确保4.24-4.30 → 5.1-5.7 → 5.8-5.14 正确排序
    periods = sorted(
        detail[PERIOD_COL].dropna().unique().tolist(),
        key=period_sort_key
    )

    # 周度店铺汇总
    weekly = (
        detail.groupby([STORE_COL, PERIOD_COL], dropna=False)
        .agg(
            GMV=("GMV", "sum"),
            订单行数=(ORDER_COL, "count"),
            订单数=(ORDER_COL, pd.Series.nunique),
            销量=(QTY_COL, "sum"),
            Base金额=(BASE_COL, "sum"),
            Shipping金额=(SHIP_COL, "sum"),
            返还VAT=("返还VAT（应收）", "sum"),
            ISR预扣=("ISR预扣（应扣）", "sum"),
            退款=(REFUND_COL, "sum"),
        )
        .reset_index()
    )
    weekly["客单价"] = weekly["GMV"] / weekly["订单数"].replace({0: pd.NA})
    weekly["件单价"] = weekly["GMV"] / weekly["销量"].replace({0: pd.NA})
    weekly = weekly.fillna(0)

    # SKU全周期汇总
    sku = (
        detail.groupby([STORE_COL, SKU_COL], dropna=False)
        .agg(
            商品名称=(PRODUCT_NAME_COL, first_non_empty),
            GMV=("GMV", "sum"),
            销量=(QTY_COL, "sum"),
            订单数=(ORDER_COL, pd.Series.nunique),
            订单行数=(ORDER_COL, "count"),
        )
        .reset_index()
    )
    sku["客单价"] = sku["GMV"] / sku["订单数"].replace({0: pd.NA})
    sku["件单价"] = sku["GMV"] / sku["销量"].replace({0: pd.NA})

    # 各周期GMV透视，用于环比计算
    pivot = (
        detail.pivot_table(index=[STORE_COL, SKU_COL], columns=PERIOD_COL, values="GMV", aggfunc="sum", fill_value=0)
        .reset_index()
    )
    for p in periods:
        if p not in pivot.columns:
            pivot[p] = 0
    sku = sku.merge(pivot[[STORE_COL, SKU_COL] + periods], on=[STORE_COL, SKU_COL], how="left")

    # 环比变化计算（适配任意数量周期）
    if len(periods) >= 2:
        sku["GMV变化"] = sku[periods[-1]] - sku[periods[-2]]
        sku["GMV变化%"] = sku["GMV变化"] / sku[periods[-2]].replace({0: pd.NA})
    else:
        sku["GMV变化"] = 0
        sku["GMV变化%"] = 0

    sku["GMV占比"] = sku["GMV"] / sku.groupby(STORE_COL)["GMV"].transform("sum").replace({0: pd.NA})
    # 修复FutureWarning，加上infer_objects
    sku = sku.fillna(0).infer_objects(copy=False).sort_values([STORE_COL, "GMV"], ascending=[True, False]).reset_index(
        drop=True)

    # SKU周度明细【修复】按周期时间排序+GMV降序
    sku_week = (
        detail.groupby([STORE_COL, PERIOD_COL, SKU_COL], dropna=False)
        .agg(
            商品名称=(PRODUCT_NAME_COL, first_non_empty),
            GMV=("GMV", "sum"),
            销量=(QTY_COL, "sum"),
            订单数=(ORDER_COL, pd.Series.nunique),
            订单行数=(ORDER_COL, "count"),
            base=(BASE_COL, "sum"),
            shipping=(SHIP_COL, "sum"),
        )
        .reset_index()
    )
    sku_week["客单价"] = sku_week["GMV"] / sku_week["订单数"].replace({0: pd.NA})
    sku_week["件单价"] = sku_week["GMV"] / sku_week["销量"].replace({0: pd.NA})
    sku_week["GMV占比"] = sku_week["GMV"] / sku_week.groupby([STORE_COL, PERIOD_COL])["GMV"].transform("sum").replace(
        {0: pd.NA})
    # 【修复】先按周期时间排序，再按GMV降序
    sku_week = sku_week.fillna(0).infer_objects(copy=False).sort_values(
        [STORE_COL, PERIOD_COL, "GMV"],
        ascending=[True, True, False],
        key=lambda x: x.map(period_sort_key) if x.name == PERIOD_COL else x
    )

    return weekly, sku, sku_week, periods


def style_header(ws, row: int, start_col: int, end_col: int):
    for cell in ws.iter_rows(min_row=row, max_row=row, min_col=start_col, max_col=end_col):
        for c in cell:
            c.fill = PatternFill("solid", fgColor=HEADER_FILL)
            c.font = Font(color=WHITE, bold=True)
            c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            c.border = BORDER


def style_table(ws, min_row: int, max_row: int, min_col: int, max_col: int):
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for c in row:
            c.border = BORDER
            c.alignment = Alignment(vertical="top", wrap_text=True)


def set_widths(ws, widths: Dict[str, float]):
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def write_rows(ws, start_row: int, start_col: int, rows: List[List]):
    for r_idx, row in enumerate(rows, start_row):
        for c_idx, value in enumerate(row, start_col):
            ws.cell(r_idx, c_idx, value)


def set_num_formats(ws, range_spec: str, fmt: str):
    for row in ws[range_spec]:
        for cell in row:
            cell.number_format = fmt


def add_top_chart(ws, top_count: int):
    if top_count <= 0:
        return
    chart = BarChart()
    chart.type = "bar"
    chart.style = 10
    chart.title = "Top SKU GMV贡献"
    chart.y_axis.title = "SKU"
    chart.x_axis.title = "GMV"
    chart.height = 10
    chart.width = 16
    data = Reference(ws, min_col=2, min_row=18, max_row=18 + top_count)
    cats = Reference(ws, min_col=1, min_row=19, max_row=18 + top_count)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.legend = None
    ws.add_chart(chart, "N18")


def build_conclusions(store: str, periods: List[str], weekly_store: pd.DataFrame, sku_store: pd.DataFrame) -> List[
    List[str]]:
    total_gmv = float(weekly_store["GMV"].sum())
    total_orders = int(weekly_store["订单数"].sum())
    total_qty = int(weekly_store["销量"].sum())
    aov = total_gmv / total_orders if total_orders else 0
    unit = total_gmv / total_qty if total_qty else 0

    rows = []
    rows.append(["总览",
                 f"全周期 GMV 为 MX${total_gmv:,.2f}，订单数 {total_orders:,}，销量 {total_qty:,}，客单价 MX${aov:,.2f}，件单价 MX${unit:,.2f}。"])

    if len(periods) >= 2:
        p1, p2 = periods[-2], periods[-1]
        w1 = weekly_store.loc[weekly_store[PERIOD_COL] == p1]
        w2 = weekly_store.loc[weekly_store[PERIOD_COL] == p2]
        if not w1.empty and not w2.empty:
            g1 = float(w1.iloc[0]["GMV"])
            g2 = float(w2.iloc[0]["GMV"])
            o1 = int(w1.iloc[0]["订单数"])
            o2 = int(w2.iloc[0]["订单数"])
            a1 = float(w1.iloc[0]["客单价"])
            a2 = float(w2.iloc[0]["客单价"])
            delta = g2 - g1
            pct = delta / g1 if g1 else 0
            dir_word = "增长" if delta >= 0 else "下降"
            rows.append(["周度变化",
                         f"{p2} GMV 为 MX${g2:,.2f}，较 {p1} {dir_word} MX${abs(delta):,.2f}（{pct:+.1%}）；订单数 {o2:,}，较上周 {o2 - o1:+,}；客单价变化 MX${a2 - a1:+,.2f}。"])

    if len(sku_store) >= 1:
        t1 = sku_store.iloc[0]
        rows.append(["第一贡献SKU",
                     f"{t1[SKU_COL]} 全周期 GMV MX${t1['GMV']:,.2f}，占比 {t1['GMV占比']:.1%}，销量 {int(t1['销量']):,}，订单 {int(t1['订单数']):,}，是最核心业绩来源。"])
    if len(sku_store) >= 2:
        t2 = sku_store.iloc[1]
        msg = f"{t2[SKU_COL]} 全周期 GMV MX${t2['GMV']:,.2f}，占比 {t2['GMV占比']:.1%}，销量 {int(t2['销量']):,}，订单 {int(t2['订单数']):,}。"
        if "GMV变化%" in sku_store.columns:
            msg += f" 最新周 GMV 环比 {t2['GMV变化%']:+.1%}。"
        rows.append(["第二贡献SKU", msg])

    if "GMV变化" in sku_store.columns and len(periods) >= 2:
        growers = sku_store[sku_store["GMV变化"] > 0].sort_values("GMV变化", ascending=False).head(3)
        if not growers.empty:
            parts = [f"{r[SKU_COL]}（+MX${r['GMV变化']:,.2f}，{r['GMV变化%']:+.1%}）" for _, r in growers.iterrows()]
            rows.append(["增长SKU", "、".join(parts) + " 最新周 GMV 增长较明显，可作为增量 SKU 关注。"])
        risks = sku_store[sku_store["GMV变化"] < 0].sort_values("GMV变化").head(3)
        if not risks.empty:
            parts = [f"{r[SKU_COL]}（-MX${abs(r['GMV变化']):,.2f}，{r['GMV变化%']:+.1%}）" for _, r in risks.iterrows()]
            rows.append(["风险SKU", "、".join(parts) + " 最新周 GMV 下滑明显，建议排查库存、价格、流量和活动变化。"])

    return rows


def create_dashboard(wb: Workbook, store: str, periods: List[str], weekly: pd.DataFrame, sku: pd.DataFrame, top_n: int):
    ws = wb.create_sheet("Dashboard")
    ws.sheet_view.showGridLines = True

    # 【修复】周度数据按时间数字排序，彻底解决顺序错乱问题
    weekly_store = weekly[weekly[STORE_COL] == store].copy().sort_values(
        PERIOD_COL,
        key=lambda x: x.map(period_sort_key)
    ).reset_index(drop=True)
    sku_store = sku[sku[STORE_COL] == store].copy().sort_values("GMV", ascending=False).reset_index(drop=True)

    period_label = "-".join([periods[0].split("-")[0], periods[-1].split("-")[-1]]) if periods else ""
    ws.merge_cells("A1:H1")
    ws["A1"] = f"{store} {period_label} GMV & SKU贡献分析"
    ws["A1"].fill = PatternFill("solid", fgColor=TITLE_FILL)
    ws["A1"].font = Font(color=WHITE, bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws.row_dimensions[1].height = 21

    total_gmv = float(weekly_store["GMV"].sum())
    total_orders = int(weekly_store["订单数"].sum())
    total_qty = int(weekly_store["销量"].sum())
    total_lines = int(weekly_store["订单行数"].sum())
    aov = total_gmv / total_orders if total_orders else 0
    unit = total_gmv / total_qty if total_qty else 0
    week_change_pct = 0
    if len(periods) >= 2:
        g1 = weekly_store.loc[weekly_store[PERIOD_COL] == periods[-2], "GMV"].sum()
        g2 = weekly_store.loc[weekly_store[PERIOD_COL] == periods[-1], "GMV"].sum()
        week_change_pct = (g2 / g1 - 1) if g1 else 0

    kpi_rows = [
        ["全周期GMV", round(total_gmv, 2)],
        ["订单数", total_orders],
        ["销量", total_qty],
        ["客单价", round(aov, 2)],
        ["件单价", round(unit, 2)],
        [f"{periods[-1] if periods else '本期'} GMV环比", week_change_pct],
    ]
    write_rows(ws, 3, 1, kpi_rows)
    style_header(ws, 3, 1, 2)
    for r in range(4, 9):
        ws.cell(r, 1).border = BORDER
        ws.cell(r, 2).border = BORDER
    ws["B8"].number_format = "0.0%"

    weekly_headers = ["周期", "GMV", "订单行数", "订单数", "销量", "Base金额", "Shipping金额", "返还VAT", "ISR预扣",
                      "退款", "客单价", "件单价"]
    write_rows(ws, 3, 4, [weekly_headers])
    weekly_rows = []
    for _, row in weekly_store.iterrows():
        weekly_rows.append([
            row[PERIOD_COL], round(row["GMV"], 2), int(row["订单行数"]), int(row["订单数"]), int(row["销量"]),
            round(row["Base金额"], 2), round(row["Shipping金额"], 2), round(row["返还VAT"], 2),
            round(row["ISR预扣"], 2),
            round(row["退款"], 2), round(row["客单价"], 2), round(row["件单价"], 2)
        ])
    write_rows(ws, 4, 4, weekly_rows)
    style_header(ws, 3, 4, 15)
    style_table(ws, 4, 3 + len(weekly_rows), 4, 15)

    ws["A10"] = "分析结论"
    ws["B10"] = "说明"
    style_header(ws, 10, 1, 2)
    conclusion_rows = build_conclusions(store, periods, weekly_store, sku_store)
    write_rows(ws, 11, 1, conclusion_rows)
    style_table(ws, 11, 10 + len(conclusion_rows), 1, 2)
    for r in range(11, 11 + len(conclusion_rows)):
        ws.cell(r, 1).font = Font(bold=True)
        ws.cell(r, 2).alignment = Alignment(wrap_text=False, vertical="center")

    top_headers = ["SKU", "GMV", "GMV占比", "销量", "订单数", "客单价", "件单价"]
    if len(periods) >= 2:
        top_headers.extend([f"{periods[-2]} GMV", f"{periods[-1]} GMV", "GMV变化", "GMV变化%"])
    else:
        top_headers.extend(["GMV变化", "GMV变化%"])
    write_rows(ws, 18, 1, [top_headers])
    top = sku_store.head(top_n).copy()
    top_rows = []
    for _, row in top.iterrows():
        base = [row[SKU_COL], round(row["GMV"], 2), row["GMV占比"], int(row["销量"]), int(row["订单数"]),
                round(row["客单价"], 2), round(row["件单价"], 2)]
        if len(periods) >= 2:
            base.extend(
                [round(row[periods[-2]], 2), round(row[periods[-1]], 2), round(row["GMV变化"], 2), row["GMV变化%"]])
        else:
            base.extend([round(row["GMV变化"], 2), row["GMV变化%"]])
        top_rows.append(base)
    write_rows(ws, 19, 1, top_rows)
    style_header(ws, 18, 1, len(top_headers))
    style_table(ws, 19, 18 + len(top_rows), 1, len(top_headers))

    # 列宽设置
    set_widths(ws, {
        "A": 16, "B": 18, "C": 13, "D": 14, "E": 13, "F": 13, "G": 13, "H": 13,
        "I": 13, "J": 13, "K": 13, "L": 13, "M": 14, "N": 13, "O": 13
    })

    # 数字格式设置
    for col in [2, 5, 9, 10, 11, 12, 13, 14, 15]:
        for row in range(3, 3 + max(6, len(weekly_rows) + 1)):
            ws.cell(row, col).number_format = "#,##0.00"
    for row in range(19, 19 + len(top_rows)):
        ws.cell(row, 2).number_format = "#,##0.00"
        ws.cell(row, 3).number_format = "0.0%"
        ws.cell(row, 6).number_format = "#,##0.00"
        ws.cell(row, 7).number_format = "#,##0.00"
        for c in range(8, min(len(top_headers), 11) + 1):
            ws.cell(row, c).number_format = "#,##0.00"
        if len(periods) >= 2:
            ws.cell(row, 11).number_format = "0.0%"
    ws["B3"].number_format = "#,##0.00"
    ws["B4"].number_format = "#,##0"
    ws["B5"].number_format = "#,##0"
    ws["B6"].number_format = "#,##0.00"
    ws["B7"].number_format = "#,##0.00"
    ws["B8"].number_format = "0.0%"

    add_top_chart(ws, len(top_rows))

    # 文本自动换行设置
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = None
    for row_cells in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=15):
        for cell in row_cells:
            current = cell.alignment
            cell.alignment = Alignment(
                horizontal=current.horizontal,
                vertical=current.vertical,
                text_rotation=current.text_rotation,
                wrap_text=False,
                shrink_to_fit=current.shrink_to_fit,
                indent=current.indent,
            )
    for header_row in [3, 10, 18]:
        for c in range(1, 16):
            if ws.cell(header_row, c).value is not None:
                ws.cell(header_row, c).alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def create_sku_total_sheet(wb: Workbook, sku: pd.DataFrame, periods: List[str]):
    ws = wb.create_sheet("SKU全周期汇总")
    headers = ["SKU", "商品名称", "GMV", "GMV占比", "销量", "订单数", "订单行数", "客单价", "件单价"]
    headers += [f"{p} GMV" for p in periods]
    headers += ["GMV变化", "GMV变化%"]
    rows = []
    for _, row in sku.sort_values([STORE_COL, "GMV"], ascending=[True, False]).iterrows():
        r = [row[SKU_COL], row["商品名称"], round(row["GMV"], 2), row["GMV占比"], int(row["销量"]), int(row["订单数"]),
             int(row["订单行数"]), round(row["客单价"], 2), round(row["件单价"], 2)]
        for p in periods:
            r.append(round(row.get(p, 0), 2))
        r += [round(row["GMV变化"], 2), row["GMV变化%"]]
        rows.append(r)
    write_rows(ws, 1, 1, [headers] + rows)
    style_header(ws, 1, 1, len(headers))
    style_table(ws, 2, 1 + len(rows), 1, len(headers))
    set_widths(ws, {"A": 16, "B": 50})
    for c in range(3, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 13
    ws.row_dimensions[1].height = 27.6
    for row in range(2, 2 + len(rows)):
        for c in [3, 8, 9] + list(range(10, 10 + len(periods))) + [10 + len(periods)]:
            ws.cell(row, c).number_format = "#,##0.00"
        ws.cell(row, 4).number_format = "0.0%"
        ws.cell(row, 11 + len(periods)).number_format = "0.0%"
    ws.freeze_panes = "A2"


def create_sku_week_sheet(wb: Workbook, sku_week: pd.DataFrame):
    ws = wb.create_sheet("SKU周度明细")
    headers = ["周期", "SKU", "商品名称", "GMV", "GMV占比", "销量", "订单数", "订单行数", "客单价", "件单价", "base",
               "shipping"]
    rows = []
    for _, row in sku_week.iterrows():
        rows.append(
            [row[PERIOD_COL], row[SKU_COL], row["商品名称"], round(row["GMV"], 2), row["GMV占比"], int(row["销量"]),
             int(row["订单数"]), int(row["订单行数"]), round(row["客单价"], 2), round(row["件单价"], 2),
             round(row["base"], 2), round(row["shipping"], 2)])
    write_rows(ws, 1, 1, [headers] + rows)
    style_header(ws, 1, 1, len(headers))
    style_table(ws, 2, 1 + len(rows), 1, len(headers))
    set_widths(ws, {"A": 13, "B": 13, "C": 50})
    for c in range(4, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 13
    for row in range(2, 2 + len(rows)):
        ws.cell(row, 4).number_format = "#,##0.00"
        ws.cell(row, 5).number_format = "0.0%"
        for c in [9, 10, 11, 12]:
            ws.cell(row, c).number_format = "#,##0.00"
    ws.freeze_panes = "A2"


def create_detail_sheet(wb: Workbook, detail: pd.DataFrame):
    ws = wb.create_sheet("GMV明细")
    headers = ["周期", "订单ID", "订单状态", "订单行ID", "订单行状态", "SKU", "商品名称", "规格", "销量",
               "Base price total", "Retail after discounts tax excl.", "Shipping tax excl.", "Product refund",
               "返还VAT（应收）", "ISR预扣（应扣）", "尾程预估（应扣）", "GMV"]
    rows = []
    for idx, row in detail.iterrows():
        excel_row = len(rows) + 2
        rows.append([
            row[PERIOD_COL], row[ORDER_COL], row[ORDER_STATUS_COL], row[ITEM_ID_COL], row[ITEM_STATUS_COL],
            row[SKU_COL], row[PRODUCT_NAME_COL], row[VARIATION_COL], int(row[QTY_COL]),
            round(row[BASE_COL], 2), round(row[RETAIL_COL], 2), round(row[SHIP_COL], 2), round(row[REFUND_COL], 2),
            f"=K{excel_row}*0.08+L{excel_row}*0.08",
            f"=-(K{excel_row}*0.025+L{excel_row}*0.025)",
            "=32",
            f"=J{excel_row}+L{excel_row}+N{excel_row}+O{excel_row}-P{excel_row}-M{excel_row}",
        ])
    write_rows(ws, 1, 1, [headers] + rows)
    style_header(ws, 1, 1, len(headers))
    style_table(ws, 2, 1 + len(rows), 1, len(headers))
    set_widths(ws, {"A": 13, "B": 22, "C": 13, "D": 22, "E": 13, "F": 16, "G": 50, "H": 13})
    for c in range(9, len(headers) + 1):
        ws.column_dimensions[get_column_letter(c)].width = 13
    ws.row_dimensions[1].height = 69
    for row in range(2, 2 + len(rows)):
        for c in range(10, 18):
            ws.cell(row, c).number_format = "#,##0.00"
    ws.freeze_panes = "A2"


def create_formula_sheet(wb: Workbook, original_rows: int, kept_rows: int, removed_rows: int, periods: List[str]):
    ws = wb.create_sheet("GMV公式说明")
    rows = [
        ["项目", "公式/口径"],
        ["返还VAT（应收）", "Retail price total after discounts(tax excl.)*0.08 + Shipping total(tax excl.)*0.08"],
        ["ISR预扣（应扣）", "-(Retail price total after discounts(tax excl.)*0.025 + Shipping total(tax excl.)*0.025)"],
        ["尾程预估（应扣）", "每个订单行固定 32"],
        ["GMV",
         "Base price total + Shipping total(tax excl.) + 返还VAT（应收） + ISR预扣（应扣） - 尾程预估（应扣） - Product refund"],
        ["数据范围",
         f"{', '.join(periods)} 订单导出文件；原始 {original_rows:,} 行，已剔除取消订单 {removed_rows:,} 行，纳入分析 {kept_rows:,} 行。"],
        ["客单价", "GMV / 去重订单数"],
        ["件单价", "GMV / 销量"],
    ]
    write_rows(ws, 1, 1, rows)
    style_header(ws, 1, 1, 2)
    style_table(ws, 2, len(rows), 1, 2)
    set_widths(ws, {"A": 22, "B": 95})
    for r in range(1, len(rows) + 1):
        ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")


def create_workbook(detail: pd.DataFrame, weekly: pd.DataFrame, sku: pd.DataFrame, sku_week: pd.DataFrame, periods: List[str], output: Path, original_rows: int, removed_rows: int, top_n: int = 15):
    wb = Workbook()
    # 自动重算Excel公式
    try:
        wb.calculation.calcMode = "auto"
        wb.calculation.fullCalcOnLoad = True
    except Exception:
        pass
    # 删除默认Sheet
    default = wb.active
    wb.remove(default)

    # 取第一个店铺作为Dashboard主店铺
    stores = sorted(detail[STORE_COL].dropna().unique().tolist())
    dashboard_store = stores[0] if stores else "店铺"
    # 筛选主店铺的周度数据，用于组合图表
    store_weekly = weekly[weekly[STORE_COL] == dashboard_store].sort_values(
        PERIOD_COL,
        key=lambda x: x.map(period_sort_key)
    ).reset_index(drop=True)

    # 1. 生成核心Dashboard看板
    create_dashboard(wb, dashboard_store, periods, weekly, sku, top_n)
    # 2. 生成其他Sheet
    create_sku_total_sheet(wb, sku, periods)
    create_sku_week_sheet(wb, sku_week)
    create_detail_sheet(wb, detail)
    create_formula_sheet(wb, original_rows, len(detail), removed_rows, periods)

    # -------------------------- 核心优化：GMV折线+销量柱形双轴组合图（全版本兼容+无标签遮挡） --------------------------
    # 1. 写入图表数据源（Dashboard空白区域，不影响原有内容）
    ws_dashboard = wb["Dashboard"]
    # 数据源位置：U列=周期，V列=GMV，W列=销量
    ws_dashboard["U1"] = "统计周期"
    ws_dashboard["V1"] = "GMV"
    ws_dashboard["W1"] = "销量"
    # 写入所有周期数据
    for idx, row in store_weekly.iterrows():
        row_num = 2 + idx
        ws_dashboard[f"U{row_num}"] = row[PERIOD_COL]
        ws_dashboard[f"V{row_num}"] = round(row["GMV"], 2)
        ws_dashboard[f"W{row_num}"] = int(row["销量"])
        # 设置数字格式
        ws_dashboard[f"V{row_num}"].number_format = "#,##0.00"
        ws_dashboard[f"W{row_num}"].number_format = "#,##0"

    # 2. 创建销量柱形图（主图表，左Y轴）
    bar_chart = BarChart()
    bar_chart.type = "col"
    bar_chart.style = 12
    bar_chart.title = f"{dashboard_store} 周度GMV&销量联动分析"
    bar_chart.y_axis.title = "销量(件)"
    bar_chart.x_axis.title = "统计周期"
    bar_chart.height = 9  # 图表高度
    bar_chart.width = 20  # 图表宽度
    # 柱形图数据源：销量
    sales_data = Reference(ws_dashboard, min_col=23, min_row=1, max_row=1 + len(store_weekly))
    categories = Reference(ws_dashboard, min_col=21, min_row=2, max_row=1 + len(store_weekly))
    bar_chart.add_data(sales_data, titles_from_data=True)
    bar_chart.set_categories(categories)
    # 柱形图样式设置
    bar_chart.series[0].graphicalProperties.solidFill = ColorChoice(srgbClr=SALES_BAR_COLOR)
    # 【核心优化】销量标签内嵌到柱形内部，不悬浮遮挡
    sales_labels = DataLabelList()
    sales_labels.showVal = True
    sales_labels.position = "inEnd"  # 柱形内部靠上位置
    bar_chart.series[0].dLbls = sales_labels
    # 关闭主Y轴网格线
    bar_chart.y_axis.majorGridlines = None

    # 3. 创建GMV折线图（次图表，右Y轴）
    line_chart = LineChart()
    line_chart.y_axis.title = "GMV总额(MX$)"
    line_chart.y_axis.axId = 200  # 次坐标轴ID
    # 折线图数据源：GMV
    gmv_data = Reference(ws_dashboard, min_col=22, min_row=1, max_row=1 + len(store_weekly))
    line_chart.add_data(gmv_data, titles_from_data=True)
    # 折线图样式设置
    line_chart.series[0].smooth = True
    line_chart.series[0].graphicalProperties.line.solidFill = ColorChoice(srgbClr=GMV_LINE_COLOR)
    line_chart.series[0].graphicalProperties.line.width = 30000  # 线宽3pt
    line_chart.series[0].marker.symbol = "circle"
    line_chart.series[0].marker.graphicalProperties.solidFill = ColorChoice(srgbClr=GMV_LINE_COLOR)
    # 【核心优化】GMV标签在点右侧，带引导线，不遮挡折线
    gmv_labels = DataLabelList()
    gmv_labels.showVal = True
    gmv_labels.position = "r"  # 数据点右侧
    gmv_labels.showLeaderLines = True  # 显示引导线，避免标签和点重叠
    line_chart.series[0].dLbls = gmv_labels
    # 关闭次Y轴网格线
    line_chart.y_axis.majorGridlines = None

    # 4. 组合图表：把折线图加到柱形图里
    bar_chart += line_chart

    # 5. 把组合图放到Dashboard的空白位置（N10开始，不影响原有内容）
    ws_dashboard.add_chart(bar_chart, "N10")
    # -----------------------------------------------------------------------------------------

    # 保存Excel文件
    wb.save(output)


def main():
    parser = argparse.ArgumentParser(description="按示例样式生成 GMV & SKU 贡献分析 Excel")
    parser.add_argument("--input", nargs="+", required=True, help="CSV 文件或包含 CSV 的文件夹")
    parser.add_argument("--output", required=True, help="输出 xlsx 文件路径")
    parser.add_argument("--top-n", type=int, default=15, help="Dashboard 展示 Top SKU 数量，默认 15")
    args = parser.parse_args()

    files = collect_csv_files(args.input)
    raw = pd.concat([load_one_csv(p) for p in files], ignore_index=True)
    original_rows = len(raw)
    filtered, removed_rows = remove_canceled_orders(raw)
    detail = calculate_detail(filtered)
    weekly, sku, sku_week, periods = build_summaries(detail)
    # 【核心修复】正确传入args.output，完全尊重命令行的输出路径设置
    create_workbook(detail, weekly, sku, sku_week, periods, Path(args.output), original_rows, removed_rows, args.top_n)

    # 打印处理结果，方便核对
    print("=" * 50)
    print(f"原始订单总行数：{original_rows:,}")
    print(f"🗑剔除取消订单数：{removed_rows:,}")
    print(f"分析有效订单：{len(detail):,}")
    print(f"全周期总GMV：{detail['GMV'].sum():,.2f} MX$")
    print(f"全周期总销量：{int(detail[QTY_COL].sum()):,} 件")
    print(f"输出文件路径：{args.output}")
    print("=" * 50)


if __name__ == "__main__":
    main()